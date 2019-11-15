import cv2
import numpy as np
import openpyxl
from matplotlib import pyplot as plt

mat=[[]]

for vid in range(37):
    skip=0
    wbkName = 'C:/Users/kedar/OneDrive - UCB-O365/Documents/Droplets/Prof. Orit/1 track motion using openCV/Book1.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    ##########################################################################################
    cap=cv2.VideoCapture('C:/Users/kedar/OneDrive - UCB-O365/Documents/Droplets/Prof. Orit/1 track motion using openCV/5D61/'+str(vid)+'.mp4')
    lx=[]
    ly=[]
    count=0
    while(cap.isOpened()):


        ret, image=cap.read()

        if ret:
            r = 300.0 / image.shape[1]
            dim =(300, int(image.shape[0] * r))

            #perform the actual resizing of the image and show it
            resized = cv2.resize(image, dim, interpolation = cv2.INTER_AREA)
            hsv = cv2.cvtColor(resized, cv2.COLOR_BGR2HSV)

            # define range of white color in HSV
            # change it according to your need !
            lower_white = np.array([30,0,0], dtype=np.uint8)
            upper_white = np.array([179,255,255], dtype=np.uint8)

            # Threshold the HSV image to get only white colors
            mask = cv2.inRange(hsv, lower_white, upper_white)
            # Bitwise-AND mask and original image
            res = cv2.bitwise_and(resized,resized, mask= mask)

            ##########################################################################################

            _,img = cv2.threshold(mask,127,255,cv2.THRESH_OTSU)
            h, w = img.shape[:2]

            contours0, hierarchy = cv2.findContours( img.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)
            moments  = [cv2.moments(cnt) for cnt in contours0]
            # Nota Bene: I rounded the centroids to integer.
            #centroids = [( int(round(m['m10']/m['m00'])),int(round(m['m01']/m['m00'])) ) for m in moments]
            for m in moments:
                if m['m00']==0:
                    continue
                else:
                    centroids = [( int(round(m['m10']/m['m00'])),int(round(m['m01']/m['m00'])) )]

            if skip%10==0:
                for m in moments:
                    if m['m00']==0:
                        continue
                    else:
                        #centroids = [( int(round(m['m10']/m['m00'])),int(round(m['m01']/m['m00'])) )]
                        count=count+1
                        if count<=10:
                            lx.append(round(m['m10']/m['m00'],2))
                            ly.append(round(m['m01']/m['m00'],2))

            skip=skip+1

            for c in centroids:
                # I draw a black little empty circle in the centroid position
                cv2.circle(img,c,5,(0,0,0),1)

            cv2.imshow('image', img)
            cv2.waitKey(1)
        if ret==0:
            break

    cap.release()
    cv2.destroyAllWindows()

    for wks in wbk.worksheets:
        for i in range((len(lx)-1)):
            wks.cell(row=i+1, column=(2*vid+1)).value = lx[i]
            wks.cell(row=i+1, column=(2*vid+2)).value = ly[i]
    if vid==0:
        mat=np.append(mat,[lx],axis=1)
        mat=np.append(mat,[ly],axis=0)
    else:
        mat=np.append(mat,[lx],axis=0)
        mat=np.append(mat,[ly],axis=0)

    wbk.save(wbkName)
    wbk.close
#mat=mat.transpose()
np.savetxt('C:/Users/kedar/OneDrive - UCB-O365/Documents/Droplets/Prof. Orit/2 angle prediction/test.txt',mat,fmt='%.2f',delimiter=' ')
